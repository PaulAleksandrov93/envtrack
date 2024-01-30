"""
Функции представлений для управления параметрами окружающей среды, комнатами,
ответственными лицами, измерительными приборами и аутентификацией пользователей.
"""

import logging
from django.utils import timezone
from django.http import HttpResponseServerError
from django.db.models import Q
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework.decorators import api_view
import xlwt
from io import BytesIO
from openpyxl import Workbook

from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView

from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from datetime import datetime, timedelta
from django.contrib.auth.models import User

from backend.models import Responsible, Room, EnviromentalParameters, MeasurementInstrument, ParameterSet
from .serializers import EnvironmentalParametersSerializer, RoomSelectSerializer, ResponsibleSerializer, MeasurementInstrumentSerializer, FilterParametersSerializer, ParameterSetSerializer


class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)

        # Add custom claims
        token['username'] = user.username
        # ...

        return token


class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getRoutes(request):
    """
    Возвращает список доступных маршрутов.

    Возвращает:
        Response: JSON-ответ с перечнем доступных маршрутов.
    """
    routes = [
        '/api/token',
        '/api/token/refresh',
    ]
    return Response(routes)


import logging

logger = logging.getLogger(__name__)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getEnviromentalParameters(request):
    try:
        """
        Возвращает все записи с параметрами окружающей среды.

        Args:
            request (Request): Объект HTTP-запроса.

        Returns:
            Response: JSON-ответ с параметрами окружающей среды.
        """
        user = request.user
        responsible = request.query_params.get('responsible')
        room = request.query_params.get('room')
        date = request.query_params.get('date')

        parameters = EnviromentalParameters.objects.all().prefetch_related('room', 'responsible', 'measurement_instrument')

        if responsible:
            parameters = parameters.filter(responsible=responsible)

        if room:
            parameters = parameters.filter(room=room)

        if date:
            created = datetime.strptime(date, '%Y-%m-%d').date()
            parameters = parameters.filter(created_at__date=created)

        parameters = parameters.order_by('-created_at')  # Добавляем сортировку по дате создания записи

        serializer = EnvironmentalParametersSerializer(parameters, many=True, context={'request': request})
        return Response(serializer.data)

    except Exception as e:
        logger.error(f'Произошла ошибка во время выполнения getEnviromentalParameters: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getEnviromentalParameter(request, pk):
    """
    Возвращает конкретную запись с параметрами окружающей среды.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ с параметрами окружающей среды.
    """
    parameters = EnviromentalParameters.objects.get(id=pk)
    serializer = EnvironmentalParametersSerializer(parameters, many=False)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getRooms(request):
    rooms = Room.objects.all()
    serializer = RoomSelectSerializer(rooms, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getMeasurementInstruments(request):
    measurement_instruments = MeasurementInstrument.objects.all()
    serializer = MeasurementInstrumentSerializer(measurement_instruments, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getResponsibles(request):
    responsibles = Responsible.objects.all()
    serializer = ResponsibleSerializer(responsibles, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createEnvironmentalParameters(request):
    print("Запрос:", request.data)
    parameter_sets_data = request.data.get('parameter_sets', [])
    parameter_set_ids = []

    for param_set_data in parameter_sets_data:
        print(f"param_set_data: {param_set_data}")  # Отладочный вывод
        parameter_set_id = param_set_data.get('id')
        print(f"Получен parameter_set_id: {parameter_set_id}")

        if parameter_set_id:
            try:
                parameter_set = ParameterSet.objects.get(id=parameter_set_id)
                parameter_set_ids.append(parameter_set.id)
                print(f"Найден ParameterSet с id {parameter_set.id}")  # Отладочный вывод
            except ParameterSet.DoesNotExist:
                return Response({'error': f'ParameterSet with id {parameter_set_id} does not exist'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            serializer = ParameterSetSerializer(data=param_set_data)
            if serializer.is_valid():
                parameter_set = serializer.save()
                parameter_set_ids.append(parameter_set.id)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    room_data = request.data.get('room')
    responsible_data = request.data.get('responsible')
    measurement_instrument_data = request.data.get('measurement_instrument')

    room, created = Room.objects.get_or_create(room_number=room_data.get('room_number'))
    print(f"Найдена комната с номером {room.room_number}")  # Отладочный вывод

    responsible, created = Responsible.objects.get_or_create(
        first_name=responsible_data.get('first_name'),
        last_name=responsible_data.get('last_name'),
        patronymic=responsible_data.get('patronymic')
    )
    print(f"Найден ответственный: {responsible.first_name} {responsible.last_name} {responsible.patronymic}")  # Отладочный вывод

    measurement_instrument, created = MeasurementInstrument.objects.get_or_create(
        serial_number=measurement_instrument_data.get('serial_number'),
        defaults=measurement_instrument_data
    )
    print(f"Найден прибор с серийным номером {measurement_instrument.serial_number}")  # Отладочный вывод

    data = {
        'room': room_data,
        'responsible': responsible_data,
        'measurement_instrument': measurement_instrument_data,
        'parameter_sets': parameter_sets_data,
        'created_at': request.data.get('created_at')
    }
    print("data=", data)
    new_serializer = EnvironmentalParametersSerializer(data=data, context={'request': request})
    if new_serializer.is_valid():
        new_serializer.save()
        print(f"Создана запись с id {new_serializer.data.get('id')}")  # Отладочный вывод
        return Response(new_serializer.data, status=status.HTTP_201_CREATED)
    else:
        print(f"Ошибка создания записи: {new_serializer.errors}")  # Отладочный вывод
        return Response(new_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateEnvironmentalParameters(request, pk):
    try:
        environmental_params = EnviromentalParameters.objects.get(pk=pk)
    except EnviromentalParameters.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = EnvironmentalParametersSerializer(instance=environmental_params, data=request.data, context={'request': request})

    if serializer.is_valid():
        room_data = request.data.get('room')
        measurement_instrument_data = request.data.get('measurement_instrument')
        parameter_sets_data = request.data.get('parameter_sets', [])
        modified_by_data = request.data.get('modified_by')
        user_id = modified_by_data.get('user')

        room, created = Room.objects.get_or_create(room_number=room_data.get('room_number')) if room_data else (None, False)
        
        measurement_instrument, created = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)

        parameter_sets = []

        for param_set_data in parameter_sets_data:
            parameter_set = ParameterSet.objects.create(
                temperature_celsius=param_set_data.get('temperature_celsius'),
                humidity_percentage=param_set_data.get('humidity_percentage'),
                pressure_kpa=param_set_data.get('pressure_kpa'),
                pressure_mmhg=param_set_data.get('pressure_mmhg'),
                time=param_set_data.get('time')
            )
            parameter_sets.append(parameter_set)

        environmental_params.room = room
        environmental_params.measurement_instrument = measurement_instrument

        environmental_params.parameter_sets.set(parameter_sets)
        
        try:
            modified_by_user = User.objects.get(id=user_id)
            environmental_params.modified_by = modified_by_user
        except User.DoesNotExist:
            print(f'Пользователь с id {user_id} не существует.')

        environmental_params.save()

        return Response(serializer.data)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteEnvironmentalParameters(request, pk):
    """
    Удаляет существующий набор параметров окружающей среды.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ параметров окружающей среды.

    Returns:
        Response: JSON-ответ, указывающий на успешное или неудачное выполнение операции.
    """
    try:
        environmental_params = EnviromentalParameters.objects.get(pk=pk)
    except EnviromentalParameters.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    environmental_params.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    """
    Получает информацию о текущем аутентифицированном пользователе.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий информацию о текущем пользователе.
    """
    user = request.user
    if user.is_authenticated:
        try:
            responsible = Responsible.objects.get(user=user)
            serializer = ResponsibleSerializer(responsible)
            return Response(serializer.data)
        except Responsible.DoesNotExist:
            return Response({'error': 'Responsible not found'}, status=404)
    else:
        return Response({'error': 'User not authenticated'}, status=401)

# ===

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getParameterSets(request):
    parameter_sets = ParameterSet.objects.all()
    serializer = ParameterSetSerializer(parameter_sets, many=True, context={'request': request})
    return Response(serializer.data)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getParameterSet(request, pk):
    parameter_set = ParameterSet.objects.get(id=pk)
    serializer = ParameterSetSerializer(parameter_set, many=False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createParameterSet(request):
    print(request.data)
    
    # Преобразовать время в формат 'HH:MM:SS'
    time_str = request.data.get('time')
    if time_str:
        try:
            datetime.strptime(time_str, '%H:%M:%S')
        except ValueError:
            return Response({'error': 'Invalid time format'}, status=status.HTTP_400_BAD_REQUEST)

    data = request.data
    if isinstance(data, list):
        created_sets = []
        for item in data:
            serializer = ParameterSetSerializer(data=item)
            print(serializer.is_valid())
            if serializer.is_valid():
                parameter_set = serializer.save()
                created_sets.append(parameter_set)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response(ParameterSetSerializer(created_sets, many=True).data, status=status.HTTP_201_CREATED)
    else:
        serializer = ParameterSetSerializer(data=data)
        if serializer.is_valid():
            parameter_set = serializer.save()
            return Response(ParameterSetSerializer(parameter_set).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateParameterSet(request, pk):
    try:
        parameter_set = ParameterSet.objects.get(pk=pk)
    except ParameterSet.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = ParameterSetSerializer(instance=parameter_set, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteParameterSet(request, pk):
    try:
        parameter_set = ParameterSet.objects.get(pk=pk)
    except ParameterSet.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    parameter_set.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
def export_parameters_to_excel(request):
    try:
        # Логика вашей функции
        logger.info(f"Export request from user: {request.user.username}")

        # Создание книги Excel
        wb = Workbook()
        ws = wb.active

        # Добавление заголовков
        columns = ['Room', 'Responsible', 'Measurement Instrument', 'Created At', 'Created By', 'Modified At', 'Modified By',
                   'Temperature (°C)', 'Humidity (%)', 'Pressure (kPa)', 'Pressure (mmHg)', 'Time']
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        parameters = EnviromentalParameters.objects.all()
        row_num = 2

        for param in parameters:
            for param_set in param.parameter_sets.all():
                param_data = [
                    param.room.room_number,
                    f'{param.responsible.last_name} {param.responsible.first_name} {param.responsible.patronymic}' if param.responsible else '',
                    param.measurement_instrument.name if param.measurement_instrument else '',
                    param.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    param.created_by.username if param.created_by else '',
                    param.modified_at.strftime('%Y-%m-%d %H:%M:%S'),
                    param.modified_by.username if param.modified_by else '',
                    param_set.temperature_celsius,
                    param_set.humidity_percentage,
                    param_set.pressure_kpa,
                    param_set.pressure_mmhg,
                    param_set.time.strftime('%H:%M:%S') if param_set.time else '',
                ]
                for col_num, param_value in enumerate(param_data, 1):
                    ws.cell(row=row_num, column=col_num, value=param_value)
                row_num += 1

        # Создание HTTP-ответа
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="environmental_parameters.xlsx"'

        # Использование BytesIO для сохранения в памяти
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response.write(buffer.read())
        buffer.close()

        return response

    except Exception as e:
        logger.error(f"Error exporting parameters: {str(e)}")
        return HttpResponseServerError("Internal Server Error")